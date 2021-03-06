from openpyxl import Workbook
wb = Workbook()
# grab the active worksheet
ws = wb.active

from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.styles.borders import Border, Side

img1 = Image('sample_image.jpg')
img1.width = 531
img1.height = 366
ws.add_image(img1, 'c5')

a = ['c','d','b','e','h','g','f']
for i in range(7):
    if i == 0 or i == 1:
        ws.column_dimensions[a[i]].width = 66.25
    elif i == 5 or i == 6:
        ws.column_dimensions[a[i]].width = 22.38
    else:
        ws.column_dimensions[a[i]].width = 1.63

for i in range(3,15):
    if i == 3 or i == 14:
        ws.row_dimensions[i].height = 13.5
    else:
        ws.row_dimensions[i].height = 91.5

ws.merge_cells('c4:d4')
ws.merge_cells('f4:g4')

top_left_cell = ws['c4']
top_left_cell.value = '사진'
top_left_cell2 = ws['f4']
top_left_cell2.value = '거래조건'

font = Font(name='경기천년제목 Bold', size=24, color='205c98')
alignment=Alignment(horizontal= 'center', vertical='center')
fillblue = PatternFill(fill_type='solid', fgColor='205c98')

for i in range(3, 15):
    ws.cell(row=i, column=2).fill = fillblue
for i in range(3, 15):
    ws.cell(row=i, column=5).fill = fillblue
for i in range(3, 15):
    ws.cell(row=i, column=8).fill = fillblue
for i in range(3, 9):
    ws.cell(row=3, column=i).fill = fillblue
for i in range(3, 9):
    ws.cell(row=4, column=i).fill = fillblue
for i in range(3, 9):
    ws.cell(row=14, column=i).fill = fillblue

ws['c4'].border = Border( left=Side(style='thin', color='ffffff'),
                          right=Side(style='thin', color='ffffff'),
                          top=Side(style='thin', color='ffffff')
                          )
ws['d4'].border = Border( left=Side(style='thin', color='ffffff'),
                          right=Side(style='thin', color='ffffff'),
                          top=Side(style='thin', color='ffffff')
                          )
ws['f4'].border = Border( left=Side(style='thin', color='ffffff'),
                          right=Side(style='thin', color='ffffff'),
                          top=Side(style='thin', color='ffffff')
                          )
ws['g4'].border = Border( left=Side(style='thin', color='ffffff'),
                          right=Side(style='thin', color='ffffff'),
                          top=Side(style='thin', color='ffffff')
                          )
for i in range(5, 14):
    ws.cell(row=i, column=2).border = Border( right=Side(style='thin', color='ffffff')
                                              )
for i in range(5, 14):
    ws.cell(row=i, column=5).border = Border( left=Side(style='thin', color='ffffff'),
                          right=Side(style='thin', color='ffffff')
                                             )
for i in range(5, 14):
    ws.cell(row=i, column=8).border = Border( left=Side(style='thin', color='ffffff')
                                             )
for i in range(3, 8):
    if i == 5:
        continue
    ws.cell(row=14, column=i).border = Border(top=Side(style='thin', color='ffffff'))

ws['c4'].font = Font(name='경기천년제목 Bold', size=36, color='ffffff')
ws['c4'].alignment = alignment

for i in range(4,14):
    if i == 4:
        ws['f4'].font = Font(name='경기천년제목 Bold', size=36, color='ffffff')
        ws['f4'].alignment = alignment
    else:
        ws.cell(row=i, column=6).font = font
        ws.cell(row=i, column=6).alignment = alignment

ws['f5'] = '층수'
ws['f6'] = '룸형태'
ws['f7'] = '크기'
ws['f8'] = '주방구조'
ws['f9'] = '보증금'
ws['f10'] = '월세'
ws['f11'] = '관리비'
ws['f12'] = '보증금조정'
ws['f13'] = '입주시기'
# Save the file
wb.save(filename='sample.xlsx')